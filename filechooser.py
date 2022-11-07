import os
import shutil
from openpyxl import load_workbook, Workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
source_folder = os.path.join(BASE_DIR,'source')
destination_folder = os.path.join(BASE_DIR,'chosen')

if not os.path.exists(source_folder):
	os.makedirs(source_folder)
if not os.path.exists(destination_folder):
	os.makedirs(destination_folder)

wb = load_workbook(filename='sample.xlsx')
ws = wb.active

wb_notfound = Workbook()
ws_notfound = wb_notfound.active
ws_notfound['A1'] = "File Not Found"
fileNotFound = False

for row in ws.rows:
	for cell in row:
		if cell.value:
			source = os.path.join(source_folder, cell.value)
			if os.path.exists(source):
				destination = os.path.join(destination_folder, cell.value)
				shutil.move(source, destination)
			else:
				ws_notfound.append([cell.value])
				fileNotFound = True

wb.close()
if fileNotFound:
	wb_notfound.save("FileNotFound.xlsx")
else:
	wb_notfound.close()